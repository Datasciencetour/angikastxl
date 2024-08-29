import streamlit as st
from transformers import AutoModelForCausalLM, AutoTokenizer
import pandas as pd
from openpyxl import load_workbook
import torch
import os

# Specify a directory to download the model to
model_dir = "angika-llm-1b"

# Check if CUDA is available and use GPU if it is
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# Download the tokenizer and model
tokenizer = AutoTokenizer.from_pretrained("satyajeet234/angika-llm-1b", cache_dir=model_dir, legacy=True)
model = AutoModelForCausalLM.from_pretrained("satyajeet234/angika-llm-1b", cache_dir=model_dir).to(device)

# Excel file path
excel_file = "generated_responses.xlsx"

# Streamlit app
st.title("Angika Language Model Chatbot")
st.write("Enter your text in Angika and see the generated output.")

# Text input from the user
input_text = st.text_input("You: ")

if st.button("Generate Response") and input_text:
    # Tokenize the input text
    input_ids = tokenizer.encode(input_text, return_tensors="pt").to(device)

    # Generate content
    outputs = model.generate(input_ids, max_length=100, num_return_sequences=1, do_sample=True, top_k=50, top_p=0.95, temperature=0.7)

    # Decode and print the generated text
    generated_text = tokenizer.decode(outputs[0], skip_special_tokens=True)

    # Display the generated text in a chatbot style
    st.write("Chatbot: " + generated_text)

    # Prepare data for Excel
    data = {"Input Text": [input_text], "Generated Text": [generated_text]}
    df = pd.DataFrame(data)

    # Write to Excel file
    if os.path.exists(excel_file):
        try:
            # Try opening the existing file to append data
            with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
                writer.book = load_workbook(excel_file)
                writer.sheets = {ws.title: ws for ws in writer.book.worksheets}
                df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row)
        except Exception as e:
            st.error("Failed to open the existing Excel file. Creating a new file.")
            df.to_excel(excel_file, index=False)  # Create a new file
    else:
        # If the file does not exist, create it and write the data with headers
        df.to_excel(excel_file, index=False)
